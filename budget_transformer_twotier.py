import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os
import logging

#Setup logging
logging.basicConfig(level=logging.INFO)
# Define the file paths
csv_file = 'your_csv_file.csv'
spreadsheet = 'your_spreadsheet.xlsx'

# Check if the CSV file exists
if not os.path.exists(csv_file):
    raise FileNotFoundError(f"CSV file not found: {csv_file}")

# Check if the spreadsheet file exists
if not os.path.exists(spreadsheet):
    raise FileNotFoundError(f"Spreadsheet file not found: {spreadsheet}")

try:
    # Load the CSV file into a DataFrame
    columns = ['Transaction Date', 'Parent Category', 'Money In', 'Money Out', 'Fee', 'Category']
    try:
        transactions_df = pd.read_csv(csv_file, usecols=columns)
    except ValueError as e:
        raise ValueError(f"Error reading CSV file: {e}")
    except Exception as e:
        raise Exception(f"An unexpected error occurred while reading the CSV file: {e}")

    # Validate the Dataframe content
    if transactions_df.empty:
        raise ValueError("The CSV file is empty or does not contain the expected columns.")
    # Check if the required columns are present in the DataFrame
    required_columns = ['Transaction Date', 'Parent Category','Money In', 'Money Out', 'Fee', 'Category']
    if not all(col in transactions_df.columns for col in required_columns):
        raise ValueError(f"The CSV file is missing one or more required columns: {required_columns}")

    # Handle missing or invalid data
    try:
        transactions_df['Money Out'] = pd.to_numeric(transactions_df['Money Out'], errors='coerce').fillna(0)
        transactions_df['Fee'] = pd.to_numeric(transactions_df['Fee'], errors='coerce').fillna(0)
    except Exception as e:
        raise ValueError(f"Error processing numeric columns: {e}")


    # Extract the month and ensure there is no missing or invalid data
    try:
        first_date = pd.to_datetime(transactions_df.at[30, 'Transaction Date'])
        month = first_date.strftime('%b')
    except KeyError:
        raise KeyError("The 'Transaction Date' column is missing or does not contain enough rows.")
    except Exception as e:
        raise ValueError(f"Error extracting month from 'Transaction Date': {e}")

    #Merge Money Out and Fees into one column
    #Convert columns to numeric, replacing any non-numeric values with 0
    transactions_df['Money Out'] = pd.to_numeric(transactions_df['Money Out'], errors='coerce').fillna(0)
    transactions_df['Fee'] = pd.to_numeric(transactions_df['Fee'], errors='coerce').fillna(0)
    #Add the columns together
    transactions_df['Money Out'] = transactions_df['Money Out'] + transactions_df['Fee']
    transactions_df.drop(columns=['Fee'], inplace=True)
    print(transactions_df)

    # Define Aggregate values by category
    category_totals = transactions_df.groupby('Category').agg({
        'Money In': 'sum',
        'Money Out': 'sum'
    }).fillna(0)
    category_totals_O = (category_totals['Money In'].fillna(0).astype(float) + category_totals['Money Out'].fillna(0).astype(float))
    print(category_totals_O)

    # Define category mapping
    # The left hand side is the category names from the Input CSV and the right hand side are names for the output spreadsheet
    category_mapping = {
        'Transfer': 'Refunds and paybacks',
        'Interest': 'Side hustle income and interest',
        'Other Income': 'Side hustle income and interest',
        #'': 'Emergencies',
        'Rent': 'Rent',
        #'': 'Subscriptions',
        'Fees': 'Bank fees',
        'Medical Aid': 'Medical aid',
        'Groceries': 'Groceries',
        #'': 'Insurance',
        'Licence': 'Transport',
        'Parking': 'Transport',
        'Vehicle Maintenance': 'Transport',
        'Public Transport': 'Transport',
        'Fuel': 'Transport',
        'Electricity': 'Electricity',
        #'': 'Water',
        'Internet': 'Internet',
        '': 'Messing around/Sus',
        'Cash Deposit': 'Side hustle income and interest',
        'Cash Withdrawal': 'Cash',
        'Education': 'Education',
        'Other Communication': 'Communication',
        'Cellphone': 'Communication',
        'Donations': 'Donations to JW',
        'Home Improvements': 'Shopping',
        'Alcohol': 'Eating out & takeaways',
        'Takeaways': 'Eating out & takeaways',
        'Restaurants': 'Eating out & takeaways',
        'Housekeeping': 'Personal & Entertainment',
        'Movies': 'Personal & Entertainment',
        'Digital Subscriptions': 'Personal & Entertainment',
        'Online Store': 'Personal & Entertainment',
        'Sport & Hobbies': 'Personal & Entertainment',
        'Gadgets': 'Personal & Entertainment',
        'Software/Games': 'Personal & Entertainment',
        'DonationsFN': 'Family',
        'Personal Care': 'Personal care',
        'Garden': 'Personal care',
        #'': 'Adventures and holidays',
        'Savings': 'Transfers between accounts',
        'Transfer': 'Transfers between accounts',
    }

    # Check for missing categories in the mapping
    for csv_category in category_totals.index:
        if csv_category not in category_mapping:
            print(f"Warning: No mapping found for category '{csv_category}'")
            logging.warning(f"No mapping found for category '{csv_category}'")

    # Load the Excel spreadsheet
    try:
        workbook = load_workbook(spreadsheet)
        sheet = workbook["Monthly Spending"]
        month_sheet = workbook["Monthly Balance Sheet"]
    except KeyError as e:
        raise KeyError(f"Error: Sheet not found in the spreadsheet: {e}")
    except Exception as e:
        raise RuntimeError(f"Error loading Excel file: {e}")

    # Find the column index for the Month
    column_found = None
    for col in month_sheet.iter_cols():
        header = col[0].value
        if isinstance(header, datetime):
            header = header.strftime('%b')
        print(header)
        if header.startswith(month):  # Check if the header starts with the month name
            column_found = col[0].column
            break
    #Validate Month Column in Excel
    if column_found is None:
        raise ValueError(f"Month {month} not found in spreadsheet")
    logging.info(f"Processing month: {month}")

    column_index = column_found - 2  # Convert to 0-based index and componsate for column offset between sheets
    logging.info(f"Found column index: {column_index}")

    # Initialize a dictionary to accumulate values for each Excel category
    excel_totals = {}

    # Accumulate values for each Excel category
    for csv_category, total_value in category_totals_O.items():
        excel_category = category_mapping.get(csv_category)
        if excel_category:
            excel_totals[excel_category] = excel_totals.get(excel_category, 0) + abs(total_value)

    # Update Excel with accumulated totals
    for row in sheet.iter_rows():
        category = row[0].value
        if category in excel_totals:
            row[column_index].value = excel_totals[category]
            #print(f"Updated {category} to {excel_totals[category]}")

    # Save the updated Excel file
    workbook.save(spreadsheet)
except Exception as e:
    print(f"An error occurred: {e}")
    logging.error(f"An error occurred: {e}")