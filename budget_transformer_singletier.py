import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# Define the file paths
csv_file = 'your_csv_file.csv'
spreadsheet = 'your_spreadsheet.xlsx'

#Each month is on a specific Column so the column for the corresponding month must be defined
column_index = 11 

# Load the CSV file into a DataFrame
columns = ['Transaction Date', 'Category', 'Money Out']
transactions_df = pd.read_csv(csv_file, usecols = columns)
print(transactions_df)

# Define Aggregate values by category
category_totals = transactions_df.groupby('Category').agg({
    'Money In': 'sum',
    'Money Out': 'sum'
}).fillna(0)
category_totals_O = category_totals['Money In'] - category_totals['Money Out']
print(category_totals_O)

# Define category mapping
# The left hand side is the category names from the Capitec Export and the right hand side are names for my category system in my spreadsheet
category_mapping = {
    'Transfer': 'Refunds and paybacks',
    'Interest': 'Side hustle income and interest',
    'Other Income': 'Side hustle income and interest',
    'Rent': 'Rent',
    'Fees': 'Bank fees',
    'Medical Aid': 'Medical aid',
    'Groceries': 'Groceries',
    'Vehicle Maintenance': 'Transport',
    'Public Transport': 'Transport',
    'Fuel': 'Transport',
    'Electricity': 'Electricity',
    'Internet': 'Internet',
    'Cash Withdrawal': 'Cash',
    'Education': 'Education',
    'Other Communication': 'Communication',
    'Cellphone': 'Communication',
    'Home Improvements': 'Shopping',
    'Alcohol': 'Eating out & takeaways',
    'Takeaways': 'Eating out & takeaways',
    'Restaurants': 'Eating out & takeaways',
    'Digital Subscriptions': 'Personal & Entertainment',
    'Going Out': 'Personal & Entertainment',
    'Sport & Hobbies': 'Personal & Entertainment',
    'DonationsFN': 'Family',
    'Housekeeping': 'Personal care',
    'Personal Care': 'Personal care',
    'Parking':'Shopping',
    'Savings': 'Transfers between accounts',
    'Transfer': 'Transfers between accounts',
}

# Load the Excel spreadsheet
workbook = load_workbook(spreadsheet)
sheet = workbook["Monthly Spending"]

# Populate the values in Excel
for csv_category, total_value in category_totals.items():
    excel_category = category_mapping.get(csv_category)
    if excel_category:
        # Find the row that matches `excel_category` and update the value
        for row in sheet.iter_rows():
            if row[0].value == excel_category:  # Assumes category names are in the first column
                row[column_index].value = abs(total_value)  # Convert negative to positive value
                break

# Save the updated Excel file
workbook.save(spreadsheet)
